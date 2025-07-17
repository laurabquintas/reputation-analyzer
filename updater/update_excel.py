from openpyxl import load_workbook
from datetime import datetime

def update_excel_rating(excel_path: str, hotel: str, source: str, rating: float):
    today = datetime.today().strftime("%Y-%m-%d")

    # Load workbook and sheet
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find the row for the given hotel
    row = None
    if hotel == "Vila Maria":
        if source == "Booking":
            row = 15

    if not row:
        print(f"❌ Hotel '{hotel}' with source '{source}' not found in Excel.")
        return

    # Find the first empty column starting from column 3
    col = 3
    while ws.cell(row=14, column=col).value:
        col += 1

    # Set date in header if new
    if not ws.cell(row=14, column=col).value:
        ws.cell(row=14, column=col).value = today

    # Set the rating
    ws.cell(row=row, column=col).value = rating

    wb.save(excel_path)
    print(f"✅ Rating {rating} for {hotel} ({source}) added on {today}")
